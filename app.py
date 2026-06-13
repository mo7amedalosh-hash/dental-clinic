import os
import io
import sys
import hashlib
import webbrowser
from functools import wraps
from datetime import datetime
from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file, jsonify
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from models import db, Patient, EditRequest, Appointment, TreatmentRecord

def get_database_uri():
    if os.environ.get('DATABASE_URL'):
        return os.environ['DATABASE_URL']
    if getattr(sys, 'frozen', False):
        db_path = os.path.join(os.path.dirname(sys.executable), 'clinic.db')
    else:
        db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'clinic.db')
    return f'sqlite:///{db_path}'

if getattr(sys, 'frozen', False):
    app = Flask(__name__, template_folder=os.path.join(sys._MEIPASS, 'templates'))
    static_folder = os.path.join(sys._MEIPASS)
else:
    app = Flask(__name__)
    static_folder = None

app.secret_key = os.environ.get('SECRET_KEY', 'dental-clinic-secret-2024')
app.config['SQLALCHEMY_DATABASE_URI'] = get_database_uri()
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)

with app.app_context():
    db.create_all()

def login_required(role=None):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not session.get('logged_in'):
                flash('الرجاء تسجيل الدخول أولاً', 'error')
                return redirect(url_for('login'))
            if role and session.get('role') != role:
                flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'error')
                return redirect(url_for('login'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

USERS = {
    'reception': {
        'password': hashlib.sha256(b'123456').hexdigest(),
        'role': 'reception',
        'display': 'الاستقبال'
    },
    'doctor': {
        'password': hashlib.sha256(b'335817').hexdigest(),
        'role': 'doctor',
        'display': 'الدكتور'
    }
}

@app.template_filter('time12')
def time12_format(time_str):
    if not time_str:
        return '—'
    try:
        parts = time_str.split(':')
        h, m = int(parts[0]), parts[1]
        ampm = 'ص' if h < 12 else 'م'
        h12 = h % 12
        if h12 == 0:
            h12 = 12
        return f'{h12}:{m} {ampm}'
    except:
        return time_str

# ─── Auth Routes ────────────────────────────────────────────────

@app.route('/ip')
def show_ip():
    try:
        import socket
        hostname = socket.gethostname()
        local_ip = socket.gethostbyname(hostname)
    except:
        hostname = 'localhost'
        local_ip = '127.0.0.1'
    return f'''<meta charset="utf-8"><h2>الاتصال من جهاز آخر</h2>
<p><b>عبر IP:</b> <a dir="ltr" href="http://{local_ip}:5000">http://{local_ip}:5000</a></p>
<p><b>عبر اسم الجهاز:</b> <a dir="ltr" href="http://{hostname}:5000">http://{hostname}:5000</a></p>
<p><b>للاستقبال:</b> <a dir="ltr" href="/reception">http://{local_ip}:5000/reception</a></p>
<p><b>للدكتور:</b> <a dir="ltr" href="/doctor">http://{local_ip}:5000/doctor</a></p>
<p style="color:green"><b>اسم الجهاز:</b> {hostname}</p>'''

@app.route('/', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        hashed = hashlib.sha256(password.encode()).hexdigest()

        if username in USERS and USERS[username]['password'] == hashed:
            session.clear()
            session['logged_in'] = True
            session['username'] = username
            session['role'] = USERS[username]['role']
            session['display'] = USERS[username]['display']

            if USERS[username]['role'] == 'doctor':
                return redirect(url_for('doctor_dashboard'))
            return redirect(url_for('reception_dashboard'))
        else:
            flash('اسم المستخدم أو كلمة المرور غير صحيحة', 'error')

    try:
        import socket
        hostname = socket.gethostname()
        local_ip = socket.gethostbyname(hostname)
    except:
        hostname = 'localhost'
        local_ip = '127.0.0.1'
    return render_template('login.html', server_ip=local_ip, server_hostname=hostname)

@app.route('/logout')
def logout():
    session.clear()
    flash('تم تسجيل الخروج بنجاح', 'info')
    return redirect(url_for('login'))

# ─── Reception Routes ──────────────────────────────────────────

@app.route('/reception')
@login_required('reception')
def reception_dashboard():
    patients = Patient.query.order_by(Patient.created_at.desc()).all()
    today = datetime.now().strftime('%Y-%m-%d')
    today_appts = Appointment.query.filter(
        Appointment.appointment_date == today,
        Appointment.status == 'scheduled'
    ).order_by(Appointment.appointment_time.asc()).all()
    return render_template('reception.html', patients=patients, today_appts=today_appts, search_query='')

@app.route('/reception/search')
@login_required('reception')
def reception_search():
    query = request.args.get('q', '').strip()
    today = datetime.now().strftime('%Y-%m-%d')
    if query:
        patients = Patient.query.filter(
            db.or_(Patient.name.like(f'%{query}%'), Patient.phone.like(f'%{query}%'))
        ).order_by(Patient.created_at.desc()).all()
    else:
        patients = Patient.query.order_by(Patient.created_at.desc()).all()
    today_appts = Appointment.query.filter(
        Appointment.appointment_date == today,
        Appointment.status == 'scheduled'
    ).order_by(Appointment.appointment_time.asc()).all()
    return render_template('reception.html', patients=patients, today_appts=today_appts, search_query=query)

@app.route('/add-patient', methods=['GET', 'POST'])
@login_required()
def add_patient():
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        if not name:
            flash('الرجاء إدخال اسم المريض', 'error')
            return render_template('add_patient.html')

        phone = request.form.get('phone', '').strip()
        address = request.form.get('address', '').strip()
        age = request.form.get('age', '').strip()
        treatment_type = request.form.get('treatment_type', '').strip()
        treatment_history = request.form.get('treatment_history', '').strip()
        doctor_name = request.form.get('doctor_name', '').strip()
        nurse_name = request.form.get('nurse_name', '').strip()
        total_fees = float(request.form.get('total_fees', 0) or 0)
        amount_paid = float(request.form.get('amount_paid', 0) or 0)

        p = Patient(
            name=name, phone=phone, address=address, age=age,
            treatment_type=treatment_type, treatment_history=treatment_history,
            doctor_name=doctor_name, nurse_name=nurse_name,
            total_fees=total_fees, amount_paid=amount_paid,
            remaining=total_fees - amount_paid,
            created_by='reception', locked=1
        )
        db.session.add(p)
        db.session.commit()
        flash(f'تم إضافة المريض {name} بنجاح', 'success')
        if session.get('role') == 'doctor':
            return redirect(url_for('doctor_dashboard'))
        return redirect(url_for('reception_dashboard'))

    return render_template('add_patient.html')

@app.route('/reception/patient/<int:patient_id>')
@login_required('reception')
def reception_patient_detail(patient_id):
    patient = db.session.get(Patient, patient_id)
    if not patient:
        flash('المريض غير موجود', 'error')
        return redirect(url_for('reception_dashboard'))
    records = TreatmentRecord.query.filter_by(patient_id=patient_id).order_by(TreatmentRecord.record_date.desc(), TreatmentRecord.created_at.desc()).all()
    return render_template('reception_patient.html', patient=patient, records=records)

@app.route('/reception/request_edit/<int:patient_id>')
@login_required('reception')
def request_edit(patient_id):
    existing = EditRequest.query.filter_by(patient_id=patient_id, status='pending').first()
    if not existing:
        patient = db.session.get(Patient, patient_id)
        if patient:
            patient.edit_requested = 1
            er = EditRequest(patient_id=patient_id, status='pending')
            db.session.add(er)
            db.session.commit()
    flash('تم إرسال طلب التعديل إلى الطبيب', 'info')
    return redirect(url_for('reception_dashboard'))

@app.route('/reception/edit/<int:patient_id>', methods=['GET', 'POST'])
@login_required('reception')
def reception_edit_patient(patient_id):
    patient = db.session.get(Patient, patient_id)
    if not patient:
        flash('المريض غير موجود', 'error')
        return redirect(url_for('reception_dashboard'))

    if patient.locked == 1:
        flash('لا يمكن التعديل. يرجى طلب الإذن من الطبيب أولاً.', 'warning')
        return redirect(url_for('reception_dashboard'))

    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        if not name:
            flash('الرجاء إدخال اسم المريض', 'error')
            return render_template('reception_edit.html', patient=patient)

        patient.name = name
        patient.phone = request.form.get('phone', '').strip()
        patient.address = request.form.get('address', '').strip()
        patient.age = request.form.get('age', '').strip()
        patient.treatment_type = request.form.get('treatment_type', '').strip()
        patient.treatment_history = request.form.get('treatment_history', '').strip()
        patient.doctor_name = request.form.get('doctor_name', '').strip()
        patient.nurse_name = request.form.get('nurse_name', '').strip()
        total_fees = float(request.form.get('total_fees', 0) or 0)
        amount_paid = float(request.form.get('amount_paid', 0) or 0)
        patient.total_fees = total_fees
        patient.amount_paid = amount_paid
        patient.remaining = total_fees - amount_paid
        patient.locked = 1
        patient.edit_requested = 0
        patient.updated_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        db.session.commit()
        flash('تم تحديث بيانات المريض بنجاح', 'success')
        return redirect(url_for('reception_dashboard'))

    return render_template('reception_edit.html', patient=patient)

# ─── Doctor Routes ─────────────────────────────────────────────

@app.route('/doctor')
@login_required('doctor')
def doctor_dashboard():
    patients = Patient.query.order_by(Patient.created_at.desc()).all()
    today = datetime.now().strftime('%Y-%m-%d')
    today_appts = Appointment.query.filter(
        Appointment.appointment_date == today,
        Appointment.status == 'scheduled'
    ).order_by(Appointment.appointment_time.asc()).all()

    edit_req_rows = EditRequest.query.filter_by(status='pending').order_by(EditRequest.requested_at.desc()).all()
    edit_requests = []
    for er in edit_req_rows:
        p = db.session.get(Patient, er.patient_id)
        edit_requests.append({
            'id': er.id, 'patient_id': er.patient_id,
            'requested_at': er.requested_at, 'status': er.status,
            'patient_name': p.name if p else ''
        })

    return render_template('doctor.html', patients=patients, edit_requests=edit_requests, today_appts=today_appts)

@app.route('/doctor/patient/<int:patient_id>')
@login_required('doctor')
def patient_card(patient_id):
    patient = db.session.get(Patient, patient_id)
    if not patient:
        flash('المريض غير موجود', 'error')
        return redirect(url_for('doctor_dashboard'))
    records = TreatmentRecord.query.filter_by(patient_id=patient_id).order_by(TreatmentRecord.record_date.desc(), TreatmentRecord.created_at.desc()).all()
    return render_template('patient_card.html', patient=patient, records=records)

@app.route('/doctor/edit/<int:patient_id>', methods=['GET', 'POST'])
@login_required('doctor')
def edit_patient(patient_id):
    patient = db.session.get(Patient, patient_id)
    if not patient:
        flash('المريض غير موجود', 'error')
        return redirect(url_for('doctor_dashboard'))

    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        if not name:
            flash('الرجاء إدخال اسم المريض', 'error')
            return render_template('edit_patient.html', patient=patient)

        patient.name = name
        patient.phone = request.form.get('phone', '').strip()
        patient.address = request.form.get('address', '').strip()
        patient.age = request.form.get('age', '').strip()
        patient.treatment_type = request.form.get('treatment_type', '').strip()
        patient.treatment_history = request.form.get('treatment_history', '').strip()
        patient.doctor_name = request.form.get('doctor_name', '').strip()
        patient.nurse_name = request.form.get('nurse_name', '').strip()
        total_fees = float(request.form.get('total_fees', 0) or 0)
        amount_paid = float(request.form.get('amount_paid', 0) or 0)
        patient.total_fees = total_fees
        patient.amount_paid = amount_paid
        patient.remaining = total_fees - amount_paid
        patient.locked = 1
        patient.edit_requested = 0
        patient.updated_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        db.session.commit()
        flash('تم تحديث بيانات المريض بنجاح', 'success')
        return redirect(url_for('patient_card', patient_id=patient_id))

    return render_template('edit_patient.html', patient=patient)

@app.route('/doctor/approve_edit/<int:patient_id>')
@login_required('doctor')
def approve_edit(patient_id):
    patient = db.session.get(Patient, patient_id)
    if patient:
        patient.locked = 0
        patient.edit_requested = 0
    EditRequest.query.filter_by(patient_id=patient_id, status='pending').update({'status': 'approved'})
    db.session.commit()
    flash('تمت الموافقة على طلب التعديل وأصبح بإمكان الاستقبال التعديل', 'success')
    return redirect(url_for('doctor_dashboard'))

@app.route('/doctor/search')
@login_required('doctor')
def doctor_search():
    query = request.args.get('q', '').strip()
    if query:
        patients = Patient.query.filter(
            db.or_(Patient.name.like(f'%{query}%'), Patient.phone.like(f'%{query}%'))
        ).order_by(Patient.created_at.desc()).all()
    else:
        patients = Patient.query.order_by(Patient.created_at.desc()).all()

    edit_req_rows = EditRequest.query.filter_by(status='pending').order_by(EditRequest.requested_at.desc()).all()
    edit_requests = []
    for er in edit_req_rows:
        p = db.session.get(Patient, er.patient_id)
        edit_requests.append({
            'id': er.id, 'patient_id': er.patient_id,
            'requested_at': er.requested_at, 'status': er.status,
            'patient_name': p.name if p else ''
        })

    return render_template('doctor.html', patients=patients, edit_requests=edit_requests, search_query=query)

@app.route('/doctor/import_excel', methods=['GET', 'POST'])
@login_required()
def import_excel():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('الرجاء اختيار ملف', 'error')
            return render_template('import_excel.html')

        file = request.files['file']
        if file.filename == '':
            flash('الرجاء اختيار ملف صالح', 'error')
            return render_template('import_excel.html')

        try:
            wb = openpyxl.load_workbook(io.BytesIO(file.read()))
            ws = wb.active
            headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]

            name_idx = phone_idx = history_idx = fees_idx = paid_idx = None
            for i, h in enumerate(headers):
                hl = h.lower()
                if any(x in hl for x in ['اسم', 'الاسم', 'name', 'patient']):
                    name_idx = i
                elif any(x in hl for x in ['هاتف', 'موبايل', 'جوال', 'phone', 'mobile']):
                    phone_idx = i
                elif any(x in hl for x in ['تاريخ', 'علاج', 'history', 'treatment']):
                    history_idx = i
                elif any(x in hl for x in ['اجمالي', 'إجمالي', 'رسوم', 'total', 'fees', 'cost']):
                    fees_idx = i
                elif any(x in hl for x in ['مدفوع', 'المدفوع', 'paid', 'payment']):
                    paid_idx = i

            if name_idx is None:
                flash('لم يتم العثور على عمود الأسماء في الملف. تأكد من وجود عمود باسم "الاسم" أو "Name"', 'error')
                return render_template('import_excel.html')

            imported = errors = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                try:
                    name = str(row[name_idx]).strip() if row[name_idx] else ''
                    if not name or name == 'None':
                        continue

                    phone = str(row[phone_idx]).strip() if phone_idx is not None and row[phone_idx] else ''
                    phone = '' if phone == 'None' else phone
                    history = str(row[history_idx]).strip() if history_idx is not None and row[history_idx] else ''
                    history = '' if history == 'None' else history
                    fees_val = row[fees_idx] if fees_idx is not None else 0
                    paid_val = row[paid_idx] if paid_idx is not None else 0
                    fees = float(fees_val) if fees_val else 0
                    paid = float(paid_val) if paid_val else 0

                    p = Patient(
                        name=name, phone=phone, treatment_history=history,
                        total_fees=fees, amount_paid=paid,
                        remaining=fees - paid, created_by='import'
                    )
                    db.session.add(p)
                    imported += 1
                except Exception:
                    errors += 1

            db.session.commit()
            msg = f'تم استيراد {imported} مريض بنجاح'
            if errors:
                msg += f' (تخطي {errors} سطر بسبب أخطاء)'
            flash(msg, 'success')
        except Exception as e:
            flash(f'خطأ في قراءة الملف: {str(e)}', 'error')

        return redirect(url_for('doctor_dashboard'))

    return render_template('import_excel.html')

# ─── Export Route ──────────────────────────────────────────────

@app.route('/export_excel')
@login_required()
def export_excel():
    patients = Patient.query.order_by(Patient.name.asc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'مرضى العيادة'

    headers = ['#', 'اسم المريض', 'العمر', 'رقم الهاتف', 'العنوان', 'نوع العلاج', 'التاريخ العلاجي', 'الطبيب', 'الممرضة', 'إجمالي الرسوم', 'المدفوع', 'المتبقي', 'ملاحظات', 'تاريخ التسجيل', 'آخر تحديث']
    header_fill = PatternFill(start_color='0D6EFD', end_color='0D6EFD', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='DEE2E6'),
        right=Side(style='thin', color='DEE2E6'),
        top=Side(style='thin', color='DEE2E6'),
        bottom=Side(style='thin', color='DEE2E6')
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    for i, p in enumerate(patients, 1):
        row = i + 1
        data = [
            i, p.name, p.age, p.phone, p.address,
            p.treatment_type, p.treatment_history,
            p.doctor_name, p.nurse_name,
            p.total_fees, p.amount_paid, p.remaining,
            p.notes, p.created_at, p.updated_at
        ]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=(col == 7 or col == 13))
            if col in (10, 11, 12):
                cell.number_format = '#,##0.00'

    for i, p in enumerate(patients):
        row = i + 2
        cell = ws.cell(row=row, column=12)
        if p.remaining == 0:
            cell.font = Font(color='198754', bold=True)
        elif p.remaining > 0:
            cell.font = Font(color='DC3545', bold=True)

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 22
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 40
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 16
    ws.column_dimensions['K'].width = 16
    ws.column_dimensions['L'].width = 16
    ws.column_dimensions['M'].width = 25
    ws.column_dimensions['N'].width = 20
    ws.column_dimensions['O'].width = 20

    ws.auto_filter.ref = f'A1:O{len(patients) + 1}'
    ws.freeze_panes = 'A2'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'مرضى_العيادة_{datetime.now().strftime("%Y-%m-%d")}.xlsx'
    )

# ─── Patient Autocomplete ──────────────────────────────────────

@app.route('/patients/search')
@login_required()
def patients_search():
    q = request.args.get('q', '').strip()
    if q:
        patients = Patient.query.filter(Patient.name.like(f'%{q}%')).order_by(Patient.name.asc()).limit(10).all()
    else:
        patients = Patient.query.order_by(Patient.name.asc()).limit(10).all()
    return jsonify([{'id': p.id, 'name': p.name, 'phone': p.phone} for p in patients])

# ─── Appointments Routes ───────────────────────────────────────

@app.route('/appointments')
@login_required()
def appointments_list():
    filter_type = request.args.get('filter', 'all')
    today = datetime.now().strftime('%Y-%m-%d')

    appt_query = Appointment.query.join(Patient, Appointment.patient_id == Patient.id).add_columns(
        Patient.name.label('patient_name'), Patient.phone.label('patient_phone')
    )

    if filter_type == 'today':
        appt_query = appt_query.filter(Appointment.appointment_date == today)
    elif filter_type == 'upcoming':
        appt_query = appt_query.filter(Appointment.appointment_date >= today, Appointment.status == 'scheduled')

    appt_rows = appt_query.order_by(Appointment.appointment_date.asc(), Appointment.appointment_time.asc()).all()
    appointments = []
    for row in appt_rows:
        a, pn, pp = row
        appointments.append({
            'id': a.id, 'patient_id': a.patient_id,
            'appointment_date': a.appointment_date,
            'appointment_time': a.appointment_time,
            'notes': a.notes, 'status': a.status,
            'created_at': a.created_at,
            'patient_name': pn, 'patient_phone': pp
        })

    return render_template('appointments.html', appointments=appointments, filter_type=filter_type, today=today)

@app.route('/appointments/add', methods=['GET', 'POST'])
@login_required()
def add_appointment():
    patients = Patient.query.with_entities(Patient.id, Patient.name, Patient.phone).order_by(Patient.name.asc()).all()

    if request.method == 'POST':
        patient_name = request.form.get('patient_name', '').strip()
        appointment_date = request.form.get('appointment_date', '').strip()
        appointment_time = request.form.get('appointment_time', '').strip()
        notes = request.form.get('notes', '').strip()

        if not patient_name or not appointment_date:
            flash('الرجاء إدخال اسم المريض وتحديد التاريخ', 'error')
            return render_template('add_appointment.html', patients=patients,
                                   patient_name=patient_name, app_date=appointment_date,
                                   app_time=appointment_time, notes=notes)

        existing = Patient.query.filter_by(name=patient_name).first()
        if existing:
            patient_id = existing.id
        else:
            p = Patient(name=patient_name, created_by='appointment')
            db.session.add(p)
            db.session.flush()
            patient_id = p.id

        a = Appointment(patient_id=patient_id, appointment_date=appointment_date,
                        appointment_time=appointment_time, notes=notes)
        db.session.add(a)
        db.session.commit()
        flash('تم إضافة الموعد بنجاح', 'success')
        return redirect(url_for('appointments_list'))

    return render_template('add_appointment.html', patients=patients,
                           patient_name=request.args.get('patient_name', ''),
                           app_date='', app_time='', notes='')

@app.route('/appointments/edit/<int:appt_id>', methods=['GET', 'POST'])
@login_required()
def edit_appointment(appt_id):
    appt = db.session.get(Appointment, appt_id)
    if not appt:
        flash('الموعد غير موجود', 'error')
        return redirect(url_for('appointments_list'))

    patient = db.session.get(Patient, appt.patient_id)
    patients = Patient.query.with_entities(Patient.id, Patient.name, Patient.phone).order_by(Patient.name.asc()).all()

    if request.method == 'POST':
        patient_name = request.form.get('patient_name', '').strip()
        appointment_date = request.form.get('appointment_date', '').strip()
        appointment_time = request.form.get('appointment_time', '').strip()
        notes = request.form.get('notes', '').strip()
        status = request.form.get('status', 'scheduled')

        if not patient_name:
            flash('الرجاء إدخال اسم المريض', 'error')
            return render_template('add_appointment.html', edit=appt,
                                   patient_name=patient_name, app_date=appointment_date,
                                   app_time=appointment_time, notes=notes, status=status)

        existing = Patient.query.filter_by(name=patient_name).first()
        if existing:
            new_patient_id = existing.id
        else:
            p = Patient(name=patient_name, created_by='appointment')
            db.session.add(p)
            db.session.flush()
            new_patient_id = p.id

        appt.patient_id = new_patient_id
        appt.appointment_date = appointment_date
        appt.appointment_time = appointment_time
        appt.notes = notes
        appt.status = status
        db.session.commit()
        flash('تم تحديث الموعد بنجاح', 'success')
        return redirect(url_for('appointments_list'))

    return render_template('add_appointment.html', edit=appt,
                           patient_name=patient.name if patient else '',
                           app_date=appt.appointment_date,
                           app_time=appt.appointment_time, notes=appt.notes,
                           status=appt.status)

@app.route('/appointments/delete/<int:appt_id>')
@login_required()
def delete_appointment(appt_id):
    appt = db.session.get(Appointment, appt_id)
    if appt:
        db.session.delete(appt)
        db.session.commit()
    flash('تم حذف الموعد', 'info')
    return redirect(url_for('appointments_list'))

@app.route('/appointments/complete/<int:appt_id>')
@login_required()
def complete_appointment(appt_id):
    appt = db.session.get(Appointment, appt_id)
    if appt:
        appt.status = 'completed'
        db.session.commit()
    flash('تم تأكيد اكتمال الموعد', 'success')
    return redirect(url_for('appointments_list'))

# ─── Treatment Records Routes ──────────────────────────────────

@app.route('/treatment/add/<int:patient_id>', methods=['POST'])
@login_required()
def add_treatment_record(patient_id):
    record_date = request.form.get('record_date', '').strip()
    description = request.form.get('description', '').strip()
    doctor_name = request.form.get('doctor_name', '').strip()
    nurse_name = request.form.get('nurse_name', '').strip()
    cost = float(request.form.get('cost', 0) or 0)

    if not record_date or not description:
        flash('الرجاء إدخال التاريخ ووصف العلاج', 'error')
    else:
        tr = TreatmentRecord(patient_id=patient_id, record_date=record_date,
                             description=description, doctor_name=doctor_name,
                             nurse_name=nurse_name, cost=cost)
        db.session.add(tr)
        total = db.session.query(db.func.coalesce(db.func.sum(TreatmentRecord.cost), 0)).filter(
            TreatmentRecord.patient_id == patient_id).scalar()
        patient = db.session.get(Patient, patient_id)
        if patient:
            patient.total_fees = total
            patient.remaining = total - patient.amount_paid
        db.session.commit()
        flash('تم إضافة السجل العلاجي', 'success')

    referrer = request.referrer or url_for('doctor_dashboard')
    return redirect(referrer)

@app.route('/treatment/delete/<int:record_id>')
@login_required('doctor')
def delete_treatment_record(record_id):
    rec = db.session.get(TreatmentRecord, record_id)
    if rec:
        patient_id = rec.patient_id
        db.session.delete(rec)
        total = db.session.query(db.func.coalesce(db.func.sum(TreatmentRecord.cost), 0)).filter(
            TreatmentRecord.patient_id == patient_id).scalar()
        patient = db.session.get(Patient, patient_id)
        if patient:
            patient.total_fees = total
            patient.remaining = total - patient.amount_paid
        db.session.commit()
    flash('تم حذف السجل العلاجي', 'info')
    return redirect(request.referrer or url_for('doctor_dashboard'))

if __name__ == '__main__':
    import socket
    hostname = socket.gethostname()
    local_ip = socket.gethostbyname(hostname)
    try:
        from waitress import serve
        print("=" * 55)
        print(f"  Hostname: {hostname}")
        print(f"  IP: {local_ip}")
        print(f"  URL: http://{local_ip}:5000")
        print(f"  Or: http://{hostname}:5000")
        print(f"  Reception: http://{local_ip}:5000/reception")
        print(f"  Doctor: http://{local_ip}:5000/doctor")
        print("=" * 55)
        webbrowser.open(f'http://127.0.0.1:5000')
        serve(app, host='0.0.0.0', port=5000)
    except ImportError:
        app.run(host='0.0.0.0', port=5000, debug=False)
